"""Tests para todos los loaders."""

import sys
import os
import pandas as pd
import tempfile
from pathlib import Path
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from src.loaders.horas import load_horas
from src.loaders.billing import load_billing
from src.loaders.presupuesto import load_presupuesto


class TestLoadHoras:
    """Tests para el loader de horas."""

    @pytest.fixture
    def sample_horas_data(self):
        """Datos de prueba para horas."""
        return {
            'Portfolio Name': ['Airlines 100Sq', 'Airlines 100Sq'],
            'Client': ['British Airways', 'British Airways'],
            'Project': ['Nexus', 'Nexus'],
            'e-Mail': ['person1@globant.com', 'person2@globant.com'],
            'Date': pd.to_datetime(['2026-06-01', '2026-06-02']),
            'Task Name': ['Task 1', 'Task 2'],
            'Hours': [8.0, 9.0],
            'Costo': [394.71, 450.00],
            'Vertical': ['Producción', 'Consultoría'],
            'Cluster': ['CODEBAY', 'CODEBAY'],
        }

    def test_load_horas_columns(self, sample_horas_data, tmp_path):
        """Test que las columnas se renombran correctamente."""
        csv_file = tmp_path / "horas.csv"
        df_input = pd.DataFrame(sample_horas_data)
        df_input.to_csv(csv_file, index=False)

        df = load_horas(str(csv_file))

        required_cols = ['cliente', 'proyecto', 'horas', 'costo']
        for col in required_cols:
            assert col in df.columns

    def test_load_horas_normalization(self, sample_horas_data, tmp_path):
        """Test que cliente y proyecto se normalizan."""
        sample_horas_data['Client'] = [' british airways ', 'BRITISH AIRWAYS']
        sample_horas_data['Project'] = ['nexus', ' NEXUS ']

        csv_file = tmp_path / "horas.csv"
        df_input = pd.DataFrame(sample_horas_data)
        df_input.to_csv(csv_file, index=False)

        df = load_horas(str(csv_file))

        assert df['cliente'].iloc[0] == 'BRITISH AIRWAYS'
        assert df['proyecto'].iloc[0] == 'NEXUS'

    def test_load_horas_types(self, sample_horas_data, tmp_path):
        """Test que los tipos de datos son correctos."""
        csv_file = tmp_path / "horas.csv"
        df_input = pd.DataFrame(sample_horas_data)
        df_input.to_csv(csv_file, index=False)

        df = load_horas(str(csv_file))

        assert df['horas'].dtype == 'float64'
        assert df['costo'].dtype == 'float64'

    def test_load_horas_from_excel(self, sample_horas_data, tmp_path):
        """Test carga desde Excel."""
        excel_file = tmp_path / "horas.xlsx"
        df_input = pd.DataFrame(sample_horas_data)
        df_input.to_excel(excel_file, index=False)

        df = load_horas(str(excel_file))

        assert df.shape[0] == 2
        assert 'cliente' in df.columns


class TestLoadBilling:
    """Tests para el loader de billing."""

    @pytest.fixture
    def sample_billing_data(self):
        """Datos de prueba para billing."""
        return {
            'Billing MONTH': pd.to_datetime(['2026-01-26', '2026-02-26']),
            'Type of Billing': ['Provision', 'Provision'],
            'Proyect': ['Fee SEO API', 'Fee SEO Web'],
            'PO': [None, None],
            'Final Billing LC': [1166.19, 1887.50],
            'Currency': ['EUR', 'EUR'],
            'Actual FX Rate': [0.85, 0.85],
            'Final Billing USD': [1371.99, 2220.59],
            'Sociedad': ['BBVA', 'BBVA'],
            'Billing number': [944868721, 944869228],
            'Actual Status': ['Approved', 'Approved'],
            'Client': ['E-BBVA', 'E-BBVA'],
            'Cliente Origen': ['BBVA S.A.', 'BBVA S.A.'],
        }

    def test_load_billing_columns(self, sample_billing_data, tmp_path):
        """Test que las columnas se renombran correctamente."""
        csv_file = tmp_path / "billing.csv"
        df_input = pd.DataFrame(sample_billing_data)
        df_input.to_csv(csv_file, index=False)

        df = load_billing(str(csv_file))

        assert 'mes' in df.columns
        assert 'cliente_origen' in df.columns
        assert 'proyecto' in df.columns
        assert 'importe_eur' in df.columns

    def test_load_billing_mes_format(self, sample_billing_data, tmp_path):
        """Test que el mes se convierte a YYYYMM."""
        csv_file = tmp_path / "billing.csv"
        df_input = pd.DataFrame(sample_billing_data)
        df_input.to_csv(csv_file, index=False)

        df = load_billing(str(csv_file))

        assert df['mes'].iloc[0] == '202601'
        assert df['mes'].iloc[1] == '202602'

    def test_load_billing_normalization(self, sample_billing_data, tmp_path):
        """Test que cliente_origen y proyecto se normalizan."""
        sample_billing_data['Cliente Origen'] = [' bbva s.a. ', 'BBVA S.A.']
        sample_billing_data['Proyect'] = ['fee seo api', ' FEE SEO WEB ']

        csv_file = tmp_path / "billing.csv"
        df_input = pd.DataFrame(sample_billing_data)
        df_input.to_csv(csv_file, index=False)

        df = load_billing(str(csv_file))

        assert df['cliente_origen'].iloc[0] == 'BBVA S.A.'
        assert df['proyecto'].iloc[0] == 'FEE SEO API'

    def test_load_billing_from_excel(self, sample_billing_data, tmp_path):
        """Test carga desde Excel."""
        excel_file = tmp_path / "billing.xlsx"
        df_input = pd.DataFrame(sample_billing_data)
        df_input.to_excel(excel_file, index=False)

        df = load_billing(str(excel_file))

        assert df.shape[0] == 2


class TestLoadPresupuesto:
    """Tests para el loader de presupuesto."""

    def test_load_presupuesto_from_real_file(self):
        """Test carga del archivo real de presupuesto."""
        filepath = "data:/Plantilla_Presupuesto_Proyectos.xlsx"
        try:
            df = load_presupuesto(filepath)

            # Verifica que contiene al menos las columnas básicas
            assert 'cliente_origen' in df.columns
            assert 'proyecto' in df.columns
            assert 'año' in df.columns

            # Verifica que tiene datos
            assert df.shape[0] > 0

            # Verifica que TO_LAUNCH es booleano si existe
            if 'to_launch' in df.columns:
                assert df['to_launch'].dtype == 'bool'
        except FileNotFoundError:
            pytest.skip("Archivo de presupuesto no encontrado")

    def test_load_presupuesto_minimal(self, tmp_path):
        """Test con presupuesto mínimo (solo datos requeridos)."""
        import openpyxl

        excel_file = tmp_path / "presupuesto.xlsx"
        wb = openpyxl.Workbook()

        # Añade instrucciones
        ws_instr = wb.active
        ws_instr.title = 'Instrucciones'
        ws_instr.append(['Instrucciones de uso...'])

        # Añade presupuesto
        ws_presup = wb.create_sheet('Presupuesto')
        ws_presup.append(['Presupuesto por proyecto'])  # Fila 1
        ws_presup.append([])  # Fila 2
        ws_presup.append([])  # Fila 3
        # Encabezados en fila 4 (índice 3)
        ws_presup.append(['Cliente Origen', 'Project', 'Año', 'Ingreso presup. (€)', 'Coste presup. (€)', 'Estado', 'TO Launch'])
        # Datos
        ws_presup.append(['CLIENT A', 'PROJECT X', 2026, 100000.0, 80000.0, 'Activo', True])
        ws_presup.append(['CLIENT B', 'PROJECT Y', 2026, 150000.0, 120000.0, 'Activo', False])

        wb.save(excel_file)

        df = load_presupuesto(str(excel_file))

        assert df.shape[0] == 2
        assert 'cliente_origen' in df.columns
        assert df['proyecto'].iloc[0] == 'PROJECT X'
        assert df['año'].iloc[0] == 2026

    def test_load_presupuesto_types(self, tmp_path):
        """Test que los tipos de datos son correctos."""
        import openpyxl

        excel_file = tmp_path / "presupuesto.xlsx"
        wb = openpyxl.Workbook()

        # Añade instrucciones
        ws_instr = wb.active
        ws_instr.title = 'Instrucciones'
        ws_instr.append(['Instrucciones'])

        # Añade presupuesto con estructura correcta
        ws_presup = wb.create_sheet('Presupuesto')
        ws_presup.append(['Presupuesto por proyecto'])
        ws_presup.append([])
        ws_presup.append([])
        # Encabezados en fila 4 (índice 3)
        ws_presup.append(['Cliente Origen', 'Project', 'Año', 'Ingreso presup. (€)', 'Coste presup. (€)', 'TO Launch'])
        # Datos
        ws_presup.append(['CLIENT A', 'PROJECT X', 2026, 100000.0, 80000.0, True])

        wb.save(excel_file)

        df = load_presupuesto(str(excel_file))

        assert df['ingreso_presupuestado'].dtype == 'float64'
        assert df['coste_presupuestado'].dtype == 'float64'
        assert df['to_launch'].dtype == 'bool'


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
